import os

import openpyxl
from app.models import University

pre = os.path.dirname(os.path.realpath(__file__))


def university_update():
    wb = openpyxl.load_workbook(pre + "/university.xlsx")
    worksheet = wb.active
    for i in range(1, worksheet.max_row):
        for col in worksheet.iter_cols(2, 2):
            if col[i].value is None:
                break
            try:
                print(col[i].value)
                name = col[i].value
                name = name.split(" «")[1][:-1]
                university = University(university=name)
                university.save()
            except Exception:
                print("ЛОХ")

        else:
            continue
        break
